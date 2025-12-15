import tkinter as tk
from tkinter import messagebox
from login import verificar_usuario
from utils import obtener_fecha_hora_actual, centrar_ventana, crear_backup
from materias_primas import abrir_materias_primas
from recetas import abrir_recetas
from base_datos import crear_tablas_recetas, RUTA_DB
from visualizar import visualizar_costos

import sqlite3
from PIL import Image, ImageTk
import os, sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# =============================
# FUNCIONES DE RECURSOS
# =============================
def resource_path(relative_path):
    """Permite acceder a recursos tanto en desarrollo como en el .exe de PyInstaller"""
    try:
        base_path = sys._MEIPASS   # carpeta temporal del .exe
    except Exception:
        base_path = os.path.abspath(".")  # carpeta del proyecto en modo .py
    return os.path.join(base_path, relative_path)


# =============================
# TABLA DE RECETAS
# =============================
def crear_tabla_recetas():
    conn = sqlite3.connect(RUTA_DB)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS recetas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT,
            peso_total REAL,
            merma REAL,
            peso_final REAL,
            costo_total REAL,
            peso_unidad REAL,
            rinde REAL,
            envase TEXT,
            costo_unidad REAL,
            fecha TEXT       
        )
    """)
    conn.commit()
    conn.close()


crear_tablas_recetas()
crear_tabla_recetas()

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def exportar_historial_openpyxl():
    conn = sqlite3.connect(RUTA_DB)
    cursor = conn.cursor()
    cursor.execute("SELECT materia, precio_anterior, precio_nuevo, fecha FROM historial_modificaciones")
    filas = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    # Encabezados con estilo
    headers = ["Materia", "Precio Anterior", "Precio Nuevo", "Fecha"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row, fila in enumerate(filas, start=2):
        for col, valor in enumerate(fila, start=1):
            ws.cell(row=row, column=col, value=valor)

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save("historial_modificaciones.xlsx")
    messagebox.showinfo("Exportación exitosa", "Historial exportado a historial_modificaciones.xlsx")


# =============================
# SISTEMA PRINCIPAL
# =============================
def iniciar_sistema(usuario):
    ventana = tk.Tk()
    ventana.title("Sistema de Costos")
    ventana.state("zoomed")
    ventana.update_idletasks()
    ventana.iconbitmap(resource_path("icono.ico"))  # Ícono de ventana

    centrar_ventana(ventana)

    # Fecha y hora en margen superior derecho
    fecha_hora = obtener_fecha_hora_actual()
    lbl_fecha = tk.Label(ventana, text=fecha_hora, anchor="e", font=("Arial", 10))
    lbl_fecha.pack(anchor="ne", padx=10, pady=5)

    # Título principal
    lbl_titulo = tk.Label(ventana, text="Sistema de Costos", font=("Arial", 28, "bold"))
    lbl_titulo.pack(pady=(30, 10))

    # =============================
    # LOGO ENTRE TÍTULO Y BIENVENIDA
    # =============================
    try:
        ruta_logo = resource_path("Images/palote_logo.jpg")
        img = Image.open(ruta_logo)
        img = img.resize((160, 90))
        logo = ImageTk.PhotoImage(img)

        frame_logo = tk.Frame(ventana)
        frame_logo.pack(pady=(0, 10))

        lbl_logo = tk.Label(frame_logo, image=logo)
        lbl_logo.image = logo  # prevenir garbage collection
        lbl_logo.pack()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar el logo: {e}", parent=ventana)

    # Bienvenida
    lbl_bienvenida = tk.Label(ventana, text=f"Bienvenido/a, {usuario}", font=("Arial", 14))
    lbl_bienvenida.pack(pady=(0, 30))

    # Frame para botones
    marco_botones = tk.Frame(ventana)
    marco_botones.pack(pady=20)

    btn_materias = tk.Button(marco_botones, text="ABM Materias Primas", width=25, height=2,
                             font=("Arial", 12), command=abrir_materias_primas)
    btn_materias.grid(row=0, column=0, padx=10, pady=10)

    btn_recetas = tk.Button(marco_botones, text="Gestión de Recetas", width=25, height=2,
                            font=("Arial", 12), command=abrir_recetas)
    btn_recetas.grid(row=1, column=0, padx=10, pady=10)

    btn_visualizar = tk.Button(marco_botones, text="Visualizar Costos", width=25, height=2,
                               font=("Arial", 12), command=visualizar_costos)
    btn_visualizar.grid(row=2, column=0, padx=10, pady=10)

    btn_exportar = tk.Button(marco_botones, text="Exportar historial", width=25, height=2,
                         font=("Arial", 12), command=exportar_historial_openpyxl)
    btn_exportar.grid(row=5, column=0, padx=10, pady=10)


    # --------------------
    # BOTÓN DE BACKUP
    # --------------------
    def boton_backup():
        ok, info = crear_backup()
        if ok:
            messagebox.showinfo("Backup creado", f"Backup guardado en:\n{info}")
        else:
            messagebox.showerror("Error", f"No se pudo crear el backup:\n{info}")

    btn_backup = tk.Button(marco_botones, text="Crear Backup", width=25, height=2,
                           font=("Arial", 12), command=boton_backup)
    btn_backup.grid(row=3, column=0, padx=10, pady=10)

    # Botón salir
    btn_salir = tk.Button(marco_botones, text="Cerrar sesión", width=25, height=2,
                          font=("Arial", 12), command=ventana.destroy)
    btn_salir.grid(row=4, column=0, padx=10, pady=10)

    # Footer
    frame_footer = tk.Frame(ventana, bg="#f0f0f0")
    frame_footer.pack(side="bottom", fill="x")

    lbl_footer = tk.Label(
        frame_footer,
        text="Sistema de Costos © 2025 Desarrollado por Mercedes Chao - Todos los derechos reservados",
        font=("Arial", 9),
        bg="#f0f0f0",
        anchor="center"
    )
    lbl_footer.pack(pady=5)

    ventana.mainloop()


# =============================
# LOGIN
# =============================
def login():
    ventana_login = tk.Tk()
    ventana_login.title("Login")
    ventana_login.geometry("400x350")
    ventana_login.iconbitmap(resource_path("icono.ico"))

    centrar_ventana(ventana_login)

    # Logo en login
    try:
        ruta_logo = resource_path("Images/palote_logo.jpg")
        img = Image.open(ruta_logo)
        img = img.resize((160, 90))
        logo = ImageTk.PhotoImage(img)

        frame_logo_login = tk.Frame(ventana_login)
        frame_logo_login.pack(pady=10)

        lbl_logo = tk.Label(frame_logo_login, image=logo)
        lbl_logo.image = logo
        lbl_logo.pack()
    except Exception as e:
        print("Error cargando logo:", e)

    tk.Label(ventana_login, text="Usuario:").pack(pady=5)
    entrada_usuario = tk.Entry(ventana_login)
    entrada_usuario.pack()

    tk.Label(ventana_login, text="Contraseña:").pack(pady=5)
    entrada_contraseña = tk.Entry(ventana_login, show="*")
    entrada_contraseña.pack()

    def validar():
        usuario = entrada_usuario.get()
        contraseña = entrada_contraseña.get()
        if verificar_usuario(usuario, contraseña):
            ventana_login.destroy()
            iniciar_sistema(usuario)
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrectos")

    tk.Button(ventana_login, text="Ingresar", command=validar).pack(pady=20)
    ventana_login.mainloop()


# =============================
# EJECUCIÓN PRINCIPAL
# =============================
if __name__ == "__main__":
    login()
